"""
FileOrganizer Pro 3.1 - Enhanced Edition
Includes Phase 1 Features: Drag & Drop, Keyboard Shortcuts, Stats Widget,
File Previews, Excel Export

Author: David - JSMS Academy
License: Proprietary
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
import shutil
import hashlib
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import json
import threading
import queue

# Import modern UI
from file_organizer_pro_modern import FileOrganizerProModern, ModernTheme

# New imports for Phase 1 features
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    DRAG_DROP_AVAILABLE = True
except ImportError:
    DRAG_DROP_AVAILABLE = False
    print("⚠️  tkinterdnd2 not installed. Drag & drop disabled.")
    print("   Install with: pip install tkinterdnd2")

try:
    from PIL import Image, ImageTk
    PREVIEW_AVAILABLE = True
except ImportError:
    PREVIEW_AVAILABLE = False
    print("⚠️  Pillow not installed. File previews disabled.")
    print("   Install with: pip install Pillow")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_EXPORT_AVAILABLE = True
except ImportError:
    EXCEL_EXPORT_AVAILABLE = False
    print("⚠️  openpyxl not installed. Excel export disabled.")
    print("   Install with: pip install openpyxl")


class FileOrganizerProV31(FileOrganizerProModern):
    """Enhanced version with Phase 1 features"""

    VERSION = "3.1.0-ENHANCED"

    def __init__(self, root):
        # Initialize with drag & drop support if available
        super().__init__(root)

        # Phase 1 features
        self.setup_drag_drop()
        self.setup_keyboard_shortcuts()
        self.thumbnail_cache = {}

        self.log("🎉 FileOrganizer Pro 3.1 Enhanced Edition loaded!")
        self.log("✨ New: Drag & Drop, Keyboard Shortcuts, Stats, Previews, Excel Export")

    def setup_drag_drop(self):
        """Feature 1: Drag & Drop Support"""
        if not DRAG_DROP_AVAILABLE:
            return

        try:
            # Enable drag and drop on main window
            self.root.drop_target_register(DND_FILES)
            self.root.dnd_bind('<<Drop>>', self.handle_drop)
            self.log("✅ Drag & Drop enabled - Drop folders onto the window!")
        except Exception as e:
            self.log(f"⚠️  Drag & Drop setup failed: {e}", "WARNING")

    def handle_drop(self, event):
        """Handle dropped files/folders"""
        try:
            # Get dropped path(s)
            files = self.root.tk.splitlist(event.data)
            if files:
                dropped_path = files[0].strip('{}')  # Remove braces if present

                # Check if it's a directory
                if os.path.isdir(dropped_path):
                    self.source_path.set(dropped_path)
                    self.log(f"📁 Dropped folder: {dropped_path}")
                    self.update_stats_widget()
                else:
                    # If it's a file, use its parent directory
                    parent_dir = os.path.dirname(dropped_path)
                    self.source_path.set(parent_dir)
                    self.log(f"📄 Dropped file: Using parent folder {parent_dir}")

        except Exception as e:
            self.log(f"Error handling drop: {e}", "ERROR")

    def setup_keyboard_shortcuts(self):
        """Feature 2: Keyboard Shortcuts"""
        shortcuts = {
            '<Control-o>': ('Browse Folder', self.browse_source),
            '<Control-s>': ('Start Organization', self.start_organization),
            '<Control-d>': ('Review Duplicates', self.review_duplicates),
            '<Control-e>': ('Manage Exclusions', self.manage_exclusions),
            '<Control-r>': ('Export Report', self.export_excel_report),
            '<F5>': ('Refresh Stats', self.update_stats_widget),
            '<Control-q>': ('Quit', lambda: self.root.quit()),
            '<Escape>': ('Stop', self.stop_organization),
        }

        for key, (desc, func) in shortcuts.items():
            self.root.bind(key, lambda e, f=func: f())

        # Log available shortcuts
        self.log("⌨️  Keyboard Shortcuts enabled:")
        for key, (desc, _) in shortcuts.items():
            shortcut_display = key.replace('<', '').replace('>', '').replace('Control', 'Ctrl')
            self.log(f"   {shortcut_display} - {desc}")

    def setup_ui(self):
        """Override setup_ui to add stats widget"""
        super().setup_ui()

        # Add stats widget to main container (find it first)
        for widget in self.root.winfo_children():
            if isinstance(widget, tk.Frame):
                main_container = widget
                break
        else:
            return

        # Insert stats widget after header (row 1)
        self.create_stats_widget(main_container)

    def create_stats_widget(self, parent):
        """Feature 3: Quick Stats Widget"""
        stats_frame = tk.LabelFrame(
            parent,
            text="  📊  Quick Statistics  ",
            bg=self.theme['bg_secondary'],
            fg=self.theme['accent_green'],
            font=('Segoe UI', 11, 'bold'),
            padx=10,
            pady=10,
            relief='flat',
            borderwidth=2,
            highlightbackground=self.theme['border'],
            highlightthickness=1
        )
        # Insert at row 1 (after header, before config)
        stats_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=10)

        # Stats container with three columns
        stats_container = tk.Frame(stats_frame, bg=self.theme['bg_secondary'])
        stats_container.pack(fill=tk.BOTH, expand=True)

        # Column 1: File counts
        col1 = tk.Frame(stats_container, bg=self.theme['bg_secondary'])
        col1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

        self.stats_labels = {}

        # Total files
        total_frame = tk.Frame(col1, bg=self.theme['bg_tertiary'], relief='flat')
        total_frame.pack(fill=tk.X, pady=3)
        tk.Label(
            total_frame,
            text="Total Files:",
            bg=self.theme['bg_tertiary'],
            fg=self.theme['text_secondary'],
            font=('Segoe UI', 9)
        ).pack(side=tk.LEFT, padx=5, pady=3)
        self.stats_labels['total'] = tk.Label(
            total_frame,
            text="0",
            bg=self.theme['bg_tertiary'],
            fg=self.theme['accent_cyan'],
            font=('Segoe UI', 11, 'bold')
        )
        self.stats_labels['total'].pack(side=tk.RIGHT, padx=5, pady=3)

        # Column 2: Size info
        col2 = tk.Frame(stats_container, bg=self.theme['bg_secondary'])
        col2.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

        # Total size
        size_frame = tk.Frame(col2, bg=self.theme['bg_tertiary'], relief='flat')
        size_frame.pack(fill=tk.X, pady=3)
        tk.Label(
            size_frame,
            text="Total Size:",
            bg=self.theme['bg_tertiary'],
            fg=self.theme['text_secondary'],
            font=('Segoe UI', 9)
        ).pack(side=tk.LEFT, padx=5, pady=3)
        self.stats_labels['size'] = tk.Label(
            size_frame,
            text="0 MB",
            bg=self.theme['bg_tertiary'],
            fg=self.theme['accent_magenta'],
            font=('Segoe UI', 11, 'bold')
        )
        self.stats_labels['size'].pack(side=tk.RIGHT, padx=5, pady=3)

        # Column 3: Category breakdown
        col3 = tk.Frame(stats_container, bg=self.theme['bg_secondary'])
        col3.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

        # Top category
        cat_frame = tk.Frame(col3, bg=self.theme['bg_tertiary'], relief='flat')
        cat_frame.pack(fill=tk.X, pady=3)
        tk.Label(
            cat_frame,
            text="Top Category:",
            bg=self.theme['bg_tertiary'],
            fg=self.theme['text_secondary'],
            font=('Segoe UI', 9)
        ).pack(side=tk.LEFT, padx=5, pady=3)
        self.stats_labels['top_cat'] = tk.Label(
            cat_frame,
            text="None",
            bg=self.theme['bg_tertiary'],
            fg=self.theme['accent_green'],
            font=('Segoe UI', 11, 'bold')
        )
        self.stats_labels['top_cat'].pack(side=tk.RIGHT, padx=5, pady=3)

        # Refresh button
        refresh_btn = tk.Button(
            stats_frame,
            text="🔄 Refresh",
            command=self.update_stats_widget,
            bg=self.theme['accent_cyan'],
            fg=self.theme['bg_primary'],
            font=('Segoe UI', 9, 'bold'),
            relief='flat',
            padx=15,
            pady=5,
            cursor='hand2'
        )
        refresh_btn.pack(pady=5)

    def update_stats_widget(self):
        """Update stats widget with current directory stats"""
        source = self.source_path.get()
        if not source or not os.path.exists(source):
            return

        self.log("📊 Scanning directory for stats...")

        # Scan in background thread
        thread = threading.Thread(target=self._scan_stats, args=(source,), daemon=True)
        thread.start()

    def _scan_stats(self, source):
        """Scan directory and update stats (runs in background)"""
        try:
            total_files = 0
            total_size = 0
            category_counts = defaultdict(int)

            for root, dirs, files in os.walk(source):
                # Skip excluded folders
                if self.is_excluded_folder(root):
                    dirs[:] = []
                    continue

                for file in files:
                    file_path = Path(root) / file
                    try:
                        total_files += 1
                        total_size += file_path.stat().st_size

                        # Categorize
                        category = self.get_file_category(file_path)
                        category_counts[category] += 1
                    except:
                        pass

            # Update UI in main thread
            def update_ui():
                self.stats_labels['total'].config(text=f"{total_files:,}")
                self.stats_labels['size'].config(text=f"{total_size / (1024**2):.1f} MB")

                if category_counts:
                    top_cat = max(category_counts.items(), key=lambda x: x[1])
                    self.stats_labels['top_cat'].config(text=f"{top_cat[0]} ({top_cat[1]})")
                else:
                    self.stats_labels['top_cat'].config(text="None")

            self.root.after(0, update_ui)
            self.root.after(0, lambda: self.log(f"✅ Stats: {total_files:,} files, {total_size / (1024**2):.1f} MB"))

        except Exception as e:
            self.root.after(0, lambda: self.log(f"Error scanning stats: {e}", "ERROR"))

    def review_duplicates(self):
        """Override to add file preview feature"""
        # Get source path
        if not self.source_path.get():
            messagebox.showwarning(
                "No Source",
                "Please select a source directory first.",
                parent=self.root
            )
            return

        recycle_bin = Path(self.source_path.get()) / "Organized" / "Duplicates_RecycleBin"
        metadata_file = recycle_bin / "duplicates_metadata.json"

        if not recycle_bin.exists() or not any(recycle_bin.iterdir()):
            messagebox.showinfo(
                "No Duplicates",
                "No duplicates found in recycle bin.",
                parent=self.root
            )
            return

        # Load metadata
        duplicates_data = {}
        if metadata_file.exists():
            try:
                with open(metadata_file, 'r', encoding='utf-8') as f:
                    duplicates_data = json.load(f)
            except (json.JSONDecodeError, IOError) as e:
                self.log(f"Warning: Could not load duplicates data: {e}", "WARNING")

        # Create dialog with preview
        self.create_duplicates_dialog_with_preview(recycle_bin, duplicates_data, metadata_file)

    def create_duplicates_dialog_with_preview(self, recycle_bin, duplicates_data, metadata_file):
        """Feature 4: File Preview in Duplicates Dialog"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Review Duplicates - With Preview")
        dialog.geometry("1100x700")
        dialog.transient(self.root)
        dialog.grab_set()

        # Header
        header_frame = tk.Frame(dialog, bg=self.theme['bg_secondary'], padx=10, pady=10)
        header_frame.pack(fill=tk.X)

        tk.Label(
            header_frame,
            text="♻️ Duplicates Recycle Bin",
            bg=self.theme['bg_secondary'],
            fg=self.theme['accent_magenta'],
            font=('Segoe UI', 16, 'bold')
        ).pack(anchor=tk.W)

        # Count duplicates
        duplicate_files = list(recycle_bin.glob('*'))
        duplicate_files = [f for f in duplicate_files if f.is_file() and f.name != 'duplicates_metadata.json']

        tk.Label(
            header_frame,
            text=f"Found {len(duplicate_files)} duplicate file(s)",
            bg=self.theme['bg_secondary'],
            fg=self.theme['text_secondary'],
            font=('Segoe UI', 10)
        ).pack(anchor=tk.W)

        # Main content area
        content_frame = tk.Frame(dialog, bg=self.theme['bg_primary'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Left: File list
        list_frame = tk.Frame(content_frame, bg=self.theme['bg_secondary'])
        list_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        tk.Label(
            list_frame,
            text="Files:",
            bg=self.theme['bg_secondary'],
            fg=self.theme['text_primary'],
            font=('Segoe UI', 10, 'bold')
        ).pack(anchor=tk.W, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            bg=self.theme['bg_tertiary'],
            fg=self.theme['text_primary'],
            font=('Consolas', 9),
            selectmode=tk.SINGLE,
            selectbackground=self.theme['accent_cyan'],
            selectforeground=self.theme['bg_primary']
        )
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.config(command=listbox.yview)

        # Populate listbox
        for file in sorted(duplicate_files, key=lambda f: f.stat().st_mtime, reverse=True):
            file_data = duplicates_data.get(file.name, {})
            moved_time = file_data.get('timestamp', file.stat().st_mtime)
            moved_date = datetime.fromtimestamp(moved_time).strftime('%Y-%m-%d')
            size_mb = file.stat().st_size / (1024 * 1024)

            display_text = f"{file.name[:40]:40} | {moved_date} | {size_mb:6.2f} MB"
            listbox.insert(tk.END, display_text)

        # Right: Preview panel
        preview_frame = tk.Frame(content_frame, bg=self.theme['bg_secondary'], width=400)
        preview_frame.pack(side=tk.RIGHT, fill=tk.BOTH, padx=(5, 0))
        preview_frame.pack_propagate(False)

        tk.Label(
            preview_frame,
            text="Preview:",
            bg=self.theme['bg_secondary'],
            fg=self.theme['text_primary'],
            font=('Segoe UI', 10, 'bold')
        ).pack(anchor=tk.W, padx=5, pady=5)

        # Preview canvas
        preview_canvas = tk.Canvas(
            preview_frame,
            bg=self.theme['bg_tertiary'],
            highlightthickness=0,
            width=380,
            height=380
        )
        preview_canvas.pack(padx=5, pady=5)

        # Details text
        details_text = scrolledtext.ScrolledText(
            preview_frame,
            height=10,
            wrap=tk.WORD,
            bg=self.theme['bg_tertiary'],
            fg=self.theme['text_primary'],
            font=('Consolas', 9)
        )
        details_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        def show_preview(event=None):
            """Show preview of selected file"""
            selection = listbox.curselection()
            if not selection:
                return

            index = selection[0]
            file = sorted(duplicate_files, key=lambda f: f.stat().st_mtime, reverse=True)[index]
            file_data = duplicates_data.get(file.name, {})

            # Clear preview
            preview_canvas.delete("all")
            details_text.delete(1.0, tk.END)

            # Show image preview if available
            if PREVIEW_AVAILABLE and file.suffix.lower() in {'.jpg', '.jpeg', '.png', '.gif', '.bmp'}:
                try:
                    img = Image.open(file)
                    img.thumbnail((360, 360))
                    photo = ImageTk.PhotoImage(img)

                    # Center image on canvas
                    x = (380 - photo.width()) // 2
                    y = (380 - photo.height()) // 2
                    preview_canvas.create_image(x, y, anchor=tk.NW, image=photo)
                    preview_canvas.image = photo  # Keep reference
                except Exception as e:
                    preview_canvas.create_text(
                        190, 190,
                        text=f"Preview unavailable\n{str(e)[:30]}",
                        fill=self.theme['text_secondary'],
                        font=('Segoe UI', 10)
                    )
            else:
                # Show file icon/type
                preview_canvas.create_text(
                    190, 190,
                    text=f"📄\n{file.suffix.upper()}",
                    fill=self.theme['accent_cyan'],
                    font=('Segoe UI', 24, 'bold')
                )

            # Show details
            details_text.insert(tk.END, f"File: {file.name}\n\n")
            details_text.insert(tk.END, f"Size: {file.stat().st_size:,} bytes\n\n")

            moved_time = file_data.get('timestamp', file.stat().st_mtime)
            details_text.insert(tk.END, f"Moved: {datetime.fromtimestamp(moved_time).strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            if 'original_path' in file_data:
                details_text.insert(tk.END, f"Original:\n{file_data['original_path']}\n\n")

            if 'md5' in file_data:
                details_text.insert(tk.END, f"MD5:\n{file_data['md5']}\n")

        listbox.bind('<<ListboxSelect>>', show_preview)

        # Action buttons
        button_frame = tk.Frame(dialog, bg=self.theme['bg_secondary'])
        button_frame.pack(fill=tk.X, padx=10, pady=10)

        def delete_selected():
            """Delete selected duplicate"""
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a file to delete")
                return

            if messagebox.askyesno("Confirm", "Permanently delete this file?"):
                index = selection[0]
                file = sorted(duplicate_files, key=lambda f: f.stat().st_mtime, reverse=True)[index]
                try:
                    file.unlink()
                    if file.name in duplicates_data:
                        del duplicates_data[file.name]
                    with open(metadata_file, 'w', encoding='utf-8') as f:
                        json.dump(duplicates_data, f, indent=2)
                    listbox.delete(index)
                    preview_canvas.delete("all")
                    details_text.delete(1.0, tk.END)
                    self.log(f"Deleted duplicate: {file.name}")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to delete: {e}")

        tk.Button(
            button_frame,
            text="🗑️ Delete Selected",
            command=delete_selected,
            bg=self.theme['error'],
            fg=self.theme['text_primary'],
            font=('Segoe UI', 10, 'bold'),
            padx=15,
            pady=8
        ).pack(side=tk.LEFT, padx=5)

        tk.Button(
            button_frame,
            text="✅ Close",
            command=dialog.destroy,
            bg=self.theme['accent_cyan'],
            fg=self.theme['bg_primary'],
            font=('Segoe UI', 10, 'bold'),
            padx=15,
            pady=8
        ).pack(side=tk.RIGHT, padx=5)

    def export_excel_report(self):
        """Feature 5: Export Report to Excel"""
        if not EXCEL_EXPORT_AVAILABLE:
            messagebox.showwarning(
                "Excel Export Unavailable",
                "openpyxl is not installed.\nInstall with: pip install openpyxl"
            )
            return

        # Check if we have stats to export
        if self.stats['files_processed'] == 0:
            messagebox.showinfo(
                "No Data",
                "Please organize files first before exporting a report."
            )
            return

        # Ask where to save
        file_path = filedialog.asksaveasfilename(
            title="Export Report",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            self.log("📊 Generating Excel report...")

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Organization Report"

            # Styling
            header_fill = PatternFill(start_color="00F7FF", end_color="00F7FF", fill_type="solid")
            header_font = Font(bold=True, size=12, color="0A0E27")

            # Title
            ws['A1'] = "FileOrganizer Pro - Organization Report"
            ws['A1'].font = Font(bold=True, size=16, color="00F7FF")
            ws.merge_cells('A1:D1')

            # Metadata
            ws['A3'] = "Generated:"
            ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws['A4'] = "Source:"
            ws['B4'] = self.source_path.get()
            ws['A5'] = "Operation:"
            ws['B5'] = self.operation_mode.get().upper()
            ws['A6'] = "Mode:"
            ws['B6'] = self.organization_mode.get()

            # Summary section
            ws['A8'] = "SUMMARY"
            ws['A8'].font = header_font
            ws['A8'].fill = header_fill
            ws.merge_cells('A8:B8')

            summary_data = [
                ("Files Processed", self.stats['files_processed']),
                ("Files Organized", self.stats['files_moved']),
                ("Duplicates Found", self.stats['duplicates_found']),
                ("Errors", self.stats['errors'])
            ]

            for i, (label, value) in enumerate(summary_data, start=9):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value
                ws[f'A{i}'].font = Font(bold=True)

            # Category breakdown
            ws['A14'] = "CATEGORY BREAKDOWN"
            ws['A14'].font = header_font
            ws['A14'].fill = header_fill
            ws.merge_cells('A14:B14')

            ws['A15'] = "Category"
            ws['B15'] = "File Count"
            ws['A15'].font = header_font
            ws['B15'].font = header_font
            ws['A15'].fill = header_fill
            ws['B15'].fill = header_fill

            row = 16
            for category, count in sorted(self.stats['categories'].items(), key=lambda x: x[1], reverse=True):
                ws[f'A{row}'] = category
                ws[f'B{row}'] = count
                row += 1

            # Adjust column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20

            # Save
            wb.save(file_path)

            self.log(f"✅ Excel report exported: {file_path}")
            messagebox.showinfo(
                "Export Complete",
                f"Report saved to:\n{file_path}"
            )

        except Exception as e:
            self.log(f"Error exporting Excel report: {e}", "ERROR")
            messagebox.showerror("Export Failed", f"Could not export report:\n{e}")


def main():
    """Main entry point for enhanced version"""

    # Use TkinterDnD root if available for drag & drop
    if DRAG_DROP_AVAILABLE:
        try:
            root = TkinterDnD.Tk()
        except:
            root = tk.Tk()
    else:
        root = tk.Tk()

    app = FileOrganizerProV31(root)
    root.mainloop()


if __name__ == "__main__":
    main()
